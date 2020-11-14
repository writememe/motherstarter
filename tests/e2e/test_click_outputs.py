# Import modules
from click.testing import CliRunner
import pytest
from motherstarter import motherstarter as ms
import traceback
import json
from openpyxl import load_workbook
import yaml

# Specify global params
base_output_dir = "motherstarter/outputs/"
base_input_dir = "tests/test_data/inputs/core/"


@pytest.fixture(scope="module")
def runner():
    return CliRunner()


def test_convert_default(runner):
    """
    Invoke motherstarter convert with defaults, so
    we can then process the various output files.
    """
    result = runner.invoke(ms.cli, ["convert"])
    if result.exception:
        traceback.print_exception(*result.exc_info)  # noqa
    expected_source_type = "DEBUG - Inventory source type is json"
    expected_output_type = "DEBUG - Output type is: all"
    """
    NOTE: Due to the default dir being a combination of the input/template dirs
    and the motherstarter_file dir, these tests are split into two.
    An example is:
    motherstarter_dir = /mnt/c/Users/acme/projects/motherstarter/
    default_template_dir = motherstarter/templates/core/
    default_source_dir = motherstarter/inputs/
    You end up with something like:
    template_dir = motherstarter_dir + default_template_dir
    source_dir = motherstarter_dir + default_source_dir
    The motherstarter_dir changes on every platform, so we don't bother
    testing that. Instead, we just want to make sure we detect the
    default_source_dir and default_template_dir in the outputs and the DEBUG stamps
    at the front of the outputs.
    """
    expected_source_dir_start = "DEBUG - Source directory is: "
    expected_source_dir_end = "motherstarter/inputs/"
    expected_template_dir_start = "DEBUG - Source template directory is: "
    expected_template_dir_end = "motherstarter/templates/core/"
    assert result.exit_code == 0
    assert expected_source_type in result.output
    assert expected_output_type in result.output
    assert expected_source_dir_start in result.output
    assert expected_source_dir_end in result.output
    assert expected_template_dir_start in result.output
    assert expected_template_dir_end in result.output


def test_output_inventory_json(
    base_input_dir=base_input_dir, base_output_dir=base_output_dir
):
    with open(f"{base_input_dir}/inventory.json") as input_f:
        input_data = json.load(input_f)
    with open(f"{base_output_dir}/json/inventory.json") as output_f:
        output_data = json.load(output_f)
    # Check that the length is the same
    assert len(output_data) == len(input_data)


def test_output_groups_json(
    base_input_dir=base_input_dir, base_output_dir=base_output_dir
):
    with open(f"{base_input_dir}/groups.json") as input_f:
        input_data = json.load(input_f)
    with open(f"{base_output_dir}/json/groups.json") as output_f:
        output_data = json.load(output_f)
    # Check that the length is the same
    assert len(output_data) == len(input_data)


def test_output_inventory_csv(
    base_input_dir=base_input_dir, base_output_dir=base_output_dir
):
    with open(f"{base_input_dir}/inventory.json") as input_f:
        input_data = json.load(input_f)
    with open(f"{base_output_dir}/csv/inventory.csv") as output_f:
        output_data = output_f.readlines()
    # Check that the length is the same
    # Minus 1 for the headers in the file
    assert len(output_data) - 1 == len(input_data)


def test_output_groups_csv(
    base_input_dir=base_input_dir, base_output_dir=base_output_dir
):
    with open(f"{base_input_dir}/groups.json") as input_f:
        input_data = json.load(input_f)
    with open(f"{base_output_dir}/csv/groups.csv") as output_f:
        output_data = output_f.readlines()
    # Check that the length is the same
    # Minus 1 for the headers in the file
    assert len(output_data) - 1 == len(input_data)


def test_output_inventory_xlsx(
    base_input_dir=base_input_dir, base_output_dir=base_output_dir
):
    with open(f"{base_input_dir}/inventory.json") as input_f:
        input_data = json.load(input_f)
    wb = load_workbook(filename=f"{base_output_dir}/xlsx/inventory.xlsx")
    ws = wb["inventory"]
    row_count = 0
    content = ws.iter_rows(
        min_row=2, min_col=0, max_row=ws.max_row, max_col=ws.max_column
    )
    # For loop to read through rows in content
    for row in content:
        row_count += 1
    # Check that the length is the same
    assert len(input_data) == row_count


def test_output_groups_xlsx(
    base_input_dir=base_input_dir, base_output_dir=base_output_dir
):
    with open(f"{base_input_dir}/groups.json") as input_f:
        input_data = json.load(input_f)
    wb = load_workbook(filename=f"{base_output_dir}/xlsx/groups.xlsx")
    ws = wb["groups"]
    row_count = 0
    content = ws.iter_rows(
        min_row=2, min_col=0, max_row=ws.max_row, max_col=ws.max_column
    )
    # For loop to read through rows in content
    for row in content:
        row_count += 1
    # Check that the length is the same
    assert len(input_data) == row_count


def test_output_inventory_nornir(
    base_input_dir=base_input_dir, base_output_dir=base_output_dir
):
    with open(f"{base_input_dir}/inventory.json") as input_f:
        input_data = json.load(input_f)
    with open(f"{base_output_dir}/nr/inventory/hosts.yaml") as output_f:
        output_data = yaml.safe_load(output_f)
    # Check that the length is the same
    assert len(output_data) == len(input_data)


def test_output_groups_nornir(
    base_input_dir=base_input_dir, base_output_dir=base_output_dir
):
    with open(f"{base_input_dir}/groups.json") as input_f:
        input_data = json.load(input_f)
    with open(f"{base_output_dir}/nr/inventory/groups.yaml") as output_f:
        output_data = yaml.safe_load(output_f)
    # Check that the length is the same
    assert len(output_data) == len(input_data)


def test_output_pyats(base_input_dir=base_input_dir, base_output_dir=base_output_dir):
    with open(f"{base_input_dir}/inventory.json") as input_f:
        input_data = json.load(input_f)
    with open(f"{base_output_dir}/pyats/mother_starter_tb.yaml") as output_f:
        output_data = yaml.safe_load(output_f)
    # Check that the length is the same
    assert len(output_data["devices"]) == len(input_data)
